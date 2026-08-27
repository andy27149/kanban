from collections import defaultdict
from pathlib import Path
import datetime
import json

import openpyxl
import pandas as pd
import yaml


# 批次分表 -> 品类/矿区代码的映射，按"计量标准"（磅单/载重）对账 合计 表 E 列每个品类的
# 入库合计，逐表核验重量取值列（标重/过磅/过磅（吨））后确定，参见项目内部对账记录。
# "5-6月A2"和"8月A3"两张表的逐行"矿区"/"发货单位"列存在 Excel 自动填充导致的递增错误
# （矿区从 A2 一路错填到 A55 等），经与用户确认：整张表按第一行的值处理为同一矿区/发货单位，
# 因此这里不逐行读取矿区/发货单位列，而是取品类为该表固定值、发货单位取第一条有效数据行的值。
BATCH_SHEET_SPECS = [
    {"sheet": "4-5月C", "category": "C", "shipper_col": 5, "weight_col": 10, "date_col": 3},
    {"sheet": "6月C", "category": "C", "shipper_col": 5, "weight_col": 10, "date_col": 3},
    {"sheet": "4-6月B", "category": "B", "shipper_col": 5, "weight_col": 9, "date_col": 3},
    {"sheet": "4月A", "category": "A", "shipper_col": 5, "weight_col": 9, "date_col": 3},
    {"sheet": "5-6月A2", "category": "A2", "shipper_col": 5, "weight_col": 9, "date_col": 3},
    {"sheet": "6月B2", "category": "B2", "shipper_col": 5, "weight_col": 10, "date_col": 3},
    {"sheet": "8月A3", "category": "A3", "shipper_col": 5, "weight_col": 9, "date_col": 3},
    {"sheet": "6-7月H", "category": "H", "shipper_col": 4, "weight_col": 11, "date_col": 7},
    {"sheet": "7月T", "category": "T", "shipper_col": 4, "weight_col": 11, "date_col": 7},
    {"sheet": "7月M", "category": "M", "shipper_col": 4, "weight_col": 11, "date_col": 7},
]
# "F" 表在同一张表里混合了 FB/FA 两个品类，靠逐行"矿区"列（第3列）区分是可靠的
# （该列只有 FB/FA 两个取值，且行数与 合计 表对账一致），发货单位固定取第一行的值。
F_SHEET_SPEC = {"sheet": "F", "region_col": 3, "shipper_col": 4, "weight_col": 7, "date_col": 2}
BATCH_HEADER_ROW = 4
BATCH_SUBHEADER_ROW = 5
BATCH_DATA_START_ROW = 6


def load_config(config_path):
    with open(config_path, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f) or {}
    return {
        "kpis": raw.get("kpis", []),
        "charts": raw.get("charts", []),
        "tables": raw.get("tables", []),
    }


def _resolve_file(uploads_dir, source_file):
    file_path = Path(uploads_dir) / source_file
    if not file_path.exists():
        raise ValueError(f"文件不存在: {source_file}")
    return file_path


def _load_worksheet(uploads_dir, source_file, sheet_name):
    file_path = _resolve_file(uploads_dir, source_file)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到工作表: {sheet_name}")
    return wb[sheet_name]


def _read_fixed_range_values(worksheet, range_str):
    result = worksheet[range_str]
    if hasattr(result, "value"):
        return [result.value]
    values = []
    for row in result:
        for cell in row:
            values.append(cell.value)
    return values


def _read_dataframe(uploads_dir, source_file, sheet_name):
    file_path = _resolve_file(uploads_dir, source_file)
    xls = pd.ExcelFile(file_path)
    if sheet_name not in xls.sheet_names:
        raise ValueError(f"找不到工作表: {sheet_name}")
    return xls.parse(sheet_name)


def _clean_value(v):
    if pd.isna(v):
        return None
    return v


def _clean_cell_value(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return v


def _read_fixed_range_column(worksheet, range_str):
    return [_clean_cell_value(v) for v in _read_fixed_range_values(worksheet, range_str)]


def _group_by_month(x_values, y_values):
    sums = {}
    for x, y in zip(x_values, y_values):
        if x is None:
            continue
        month = str(x)[:7]
        sums[month] = sums.get(month, 0) + (y or 0)
    months = sorted(sums.keys())
    return months, [sums[m] for m in months]


def _read_sheet_rows(uploads_dir, source_file, sheet_name, header_row, id_column):
    """按行读取整张 sheet（不依赖 pandas，允许重复列名）。
    header_row 之上的行（如标题行）被忽略；id_column 所在列为空的行自动跳过。
    """
    worksheet = _load_worksheet(uploads_dir, source_file, sheet_name)
    header_cells = next(worksheet.iter_rows(min_row=header_row, max_row=header_row))
    columns = [_clean_cell_value(cell.value) for cell in header_cells]
    try:
        id_index = columns.index(id_column)
    except ValueError:
        raise ValueError(f"找不到列: {id_column}")

    rows = []
    for row_cells in worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row):
        values = [_clean_cell_value(cell.value) for cell in row_cells[: len(columns)]]
        if values[id_index] is None:
            continue
        rows.append(values)
    return columns, rows


def _build_totals_row(columns, rows, id_column):
    id_index = columns.index(id_column)
    totals = [None] * len(columns)
    totals[id_index] = f"合计（{len(rows)} 份合同）"
    for col_index in range(len(columns)):
        if col_index == id_index:
            continue
        non_null = [row[col_index] for row in rows if row[col_index] is not None]
        if non_null and all(isinstance(v, (int, float)) for v in non_null):
            totals[col_index] = sum(non_null)
    return totals


def _sum_range(worksheet, range_str):
    values = _read_fixed_range_values(worksheet, range_str)
    numeric_values = [v for v in values if isinstance(v, (int, float))]
    return sum(numeric_values) if numeric_values else 0


def _sheet_headers(worksheet, header_row, sub_row):
    headers = []
    last_top = None
    for c in range(1, worksheet.max_column + 1):
        top = worksheet.cell(row=header_row, column=c).value
        if top is not None:
            last_top = top
        sub = worksheet.cell(row=sub_row, column=c).value
        if top and sub:
            headers.append(f"{top}-{sub}")
        elif sub:
            headers.append(f"{last_top}-{sub}" if last_top else sub)
        elif top:
            headers.append(top)
        else:
            headers.append(f"列{c}")
    return headers


def _scan_batch_records(uploads_dir, source_file):
    file_path = _resolve_file(uploads_dir, source_file)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    records = []

    def first_shipper(ws, shipper_col):
        for r in range(BATCH_DATA_START_ROW, ws.max_row + 1):
            v = ws.cell(row=r, column=shipper_col).value
            if v is not None:
                return v
        return None

    def row_values(ws, r):
        return [_clean_cell_value(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]

    for spec in BATCH_SHEET_SPECS:
        if spec["sheet"] not in wb.sheetnames:
            continue
        ws = wb[spec["sheet"]]
        headers = _sheet_headers(ws, BATCH_HEADER_ROW, BATCH_SUBHEADER_ROW)
        shipper = first_shipper(ws, spec["shipper_col"])
        for r in range(BATCH_DATA_START_ROW, ws.max_row + 1):
            weight = ws.cell(row=r, column=spec["weight_col"]).value
            date_val = _clean_cell_value(ws.cell(row=r, column=spec["date_col"]).value)
            # 每张批次表最后一行是不带日期的"合计"行（重量列会重复汇总值），
            # 必须靠日期非空来排除，否则会把合计行当成一条记录导致重量翻倍。
            if not isinstance(weight, (int, float)) or date_val is None:
                continue
            records.append({
                "category": spec["category"],
                "shipper": shipper,
                "weight": weight,
                "date": date_val,
                "sheet": spec["sheet"],
                "headers": headers,
                "row": row_values(ws, r),
            })

    if F_SHEET_SPEC["sheet"] in wb.sheetnames:
        spec = F_SHEET_SPEC
        ws = wb[spec["sheet"]]
        headers = _sheet_headers(ws, BATCH_HEADER_ROW, BATCH_SUBHEADER_ROW)
        shipper = first_shipper(ws, spec["shipper_col"])
        for r in range(BATCH_DATA_START_ROW, ws.max_row + 1):
            region = ws.cell(row=r, column=spec["region_col"]).value
            weight = ws.cell(row=r, column=spec["weight_col"]).value
            if region is None or not isinstance(weight, (int, float)):
                continue
            date_val = _clean_cell_value(ws.cell(row=r, column=spec["date_col"]).value)
            records.append({
                "category": region,
                "shipper": shipper,
                "weight": weight,
                "date": date_val,
                "sheet": spec["sheet"],
                "headers": headers,
                "row": row_values(ws, r),
            })

    return records


def build_category_sources(records):
    by_category = defaultdict(lambda: defaultdict(list))
    headers_by_category_sheet = {}
    for rec in records:
        by_category[rec["category"]][rec["sheet"]].append(rec["row"])
        headers_by_category_sheet[(rec["category"], rec["sheet"])] = rec["headers"]

    result = {}
    for category, sheets in by_category.items():
        result[category] = [
            {"sheet": sheet, "headers": headers_by_category_sheet[(category, sheet)], "rows": rows}
            for sheet, rows in sheets.items()
        ]
    return result


def _read_category_inventory(uploads_dir, source_file):
    worksheet = _load_worksheet(uploads_dir, source_file, "合计")
    categories = _read_fixed_range_column(worksheet, "B6:B16")
    inventories = _read_fixed_range_column(worksheet, "K6:K16")
    return {cat: (inv or 0) for cat, inv in zip(categories, inventories) if cat is not None}


def build_shipper_summary(records, category_inventory):
    shipper_category_weight = defaultdict(lambda: defaultdict(float))
    shipper_month_weight = defaultdict(lambda: defaultdict(float))
    category_total_weight = defaultdict(float)

    for rec in records:
        shipper_category_weight[rec["shipper"]][rec["category"]] += rec["weight"]
        category_total_weight[rec["category"]] += rec["weight"]
        if rec["date"]:
            month = str(rec["date"])[:7]
            shipper_month_weight[rec["shipper"]][month] += rec["weight"]

    shippers = sorted(shipper_category_weight.keys())

    by_shipper_rows = []
    for shipper in shippers:
        cat_weights = shipper_category_weight[shipper]
        total_weight = sum(cat_weights.values())
        inventory_share = 0.0
        for category, w in cat_weights.items():
            cat_total = category_total_weight.get(category) or 0
            cat_inv = category_inventory.get(category) or 0
            if cat_total:
                inventory_share += cat_inv * (w / cat_total)
        by_shipper_rows.append([shipper, round(total_weight, 2), round(inventory_share, 2)])

    all_months = sorted({m for months in shipper_month_weight.values() for m in months})
    monthly_rows = []
    for shipper in shippers:
        row = [shipper] + [round(shipper_month_weight[shipper].get(m, 0.0), 2) for m in all_months]
        monthly_rows.append(row)

    return {
        "by_shipper": {"columns": ["发货单位", "发运总量（吨）", "分摊库存余额（吨）"], "rows": by_shipper_rows},
        "by_shipper_monthly": {"columns": ["发货单位"] + all_months, "rows": monthly_rows},
    }


def _detail_group_column_range(worksheet, group_row, group_label):
    for rng in worksheet.merged_cells.ranges:
        if rng.min_row == group_row and rng.max_row == group_row:
            if worksheet.cell(row=rng.min_row, column=rng.min_col).value == group_label:
                return rng.min_col, rng.max_col
    for c in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=group_row, column=c).value == group_label:
            return c, c
    return None, None


def _merged_cell_value(worksheet, row, col):
    for rng in worksheet.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return worksheet.cell(row=rng.min_row, column=rng.min_col).value
    return worksheet.cell(row=row, column=col).value


def _extract_detail_group(worksheet, spec):
    start_col, end_col = _detail_group_column_range(worksheet, spec["group_label_row"], spec["group_label"])
    if start_col is None:
        return []

    labels = {}
    for c in range(start_col, end_col + 1):
        name = _merged_cell_value(worksheet, spec["category_row"], c)
        metric = _merged_cell_value(worksheet, spec["metric_row"], c)
        if name and metric:
            labels[c] = f"{name}-{metric}"
        elif name:
            labels[c] = name
        elif metric:
            labels[c] = metric

    row_extras = []
    for r in range(spec["data_start_row"], spec["data_end_row"] + 1):
        items = []
        for c, label in labels.items():
            v = worksheet.cell(row=r, column=c).value
            if v is not None:
                items.append({"label": label, "value": _clean_cell_value(v)})
        row_extras.append(items)
    return row_extras


def extract_kpi(item, uploads_dir):
    key = item["key"]
    label = item["label"]
    group_fields = {
        k: item[k] for k in ("group", "group_icon", "group_label") if k in item
    }
    try:
        mode = item["mode"]
        if mode == "fixed_range":
            worksheet = _load_worksheet(uploads_dir, item["source_file"], item["sheet"])
            values = _read_fixed_range_values(worksheet, item["range"])
            numeric_values = [v for v in values if isinstance(v, (int, float))]
            value = sum(numeric_values) if numeric_values else None
        elif mode == "header_match":
            df = _read_dataframe(uploads_dir, item["source_file"], item["sheet"])
            header = item["header"]
            if header not in df.columns:
                raise ValueError(f"找不到表头: '{header}'")
            values = [_clean_value(v) for v in df[header].tolist()]
            numeric_values = [v for v in values if isinstance(v, (int, float))]
            value = sum(numeric_values) if numeric_values else None
        elif mode == "header_sum":
            columns, rows = _read_sheet_rows(
                uploads_dir, item["source_file"], item["sheet"], item["header_row"], item["id_column"]
            )
            if not rows:
                value = None
            else:
                indices = []
                for col in item["value_columns"]:
                    try:
                        indices.append(columns.index(col))
                    except ValueError:
                        raise ValueError(f"找不到列: {col}")
                value = sum(
                    row[idx]
                    for row in rows
                    for idx in indices
                    if isinstance(row[idx], (int, float))
                )
        else:
            raise ValueError(f"未知的取数模式: {mode}")
        return {"key": key, "label": label, "value": value, "error": None, **group_fields}
    except Exception as exc:
        return {"key": key, "label": label, "value": None, "error": str(exc), **group_fields}


def extract_chart(item, uploads_dir):
    key = item["key"]
    try:
        mode = item.get("mode", "header_match")
        if mode == "fixed_range":
            worksheet = _load_worksheet(uploads_dir, item["source_file"], item["sheet"])
            x = _read_fixed_range_column(worksheet, item["x_range"])
            result = {
                "key": key,
                "type": item["type"],
                "title": item["title"],
                "x": x,
                "error": None,
            }
            if "series" in item:
                result["series"] = [
                    {"name": s["name"], "data": _read_fixed_range_column(worksheet, s["range"])}
                    for s in item["series"]
                ]
            else:
                y = _read_fixed_range_column(worksheet, item["y_range"])
                if item.get("group_by") == "month":
                    x, y = _group_by_month(x, y)
                    result["x"] = x
                result["y"] = y
            return result
        elif mode == "group_by":
            columns, rows = _read_sheet_rows(
                uploads_dir, item["source_file"], item["sheet"], item["header_row"], item["id_column"]
            )
            group_column = item["group_column"]
            try:
                group_index = columns.index(group_column)
            except ValueError:
                raise ValueError(f"找不到列: {group_column}")
            agg = item.get("agg", "count")
            sums = {}
            if agg == "sum":
                value_column = item["value_column"]
                try:
                    value_index = columns.index(value_column)
                except ValueError:
                    raise ValueError(f"找不到列: {value_column}")
                for row in rows:
                    group_value = row[group_index]
                    if group_value is None:
                        continue
                    cell_value = row[value_index]
                    numeric_value = cell_value if isinstance(cell_value, (int, float)) else 0
                    sums[group_value] = sums.get(group_value, 0) + numeric_value
            else:
                for row in rows:
                    group_value = row[group_index]
                    if group_value is None:
                        continue
                    sums[group_value] = sums.get(group_value, 0) + 1
            x = list(sums.keys())
            y = [sums[k] for k in x]
            return {
                "key": key,
                "type": item["type"],
                "title": item["title"],
                "x": x,
                "y": y,
                "error": None,
            }
        elif mode == "sum_bars":
            worksheet = _load_worksheet(uploads_dir, item["source_file"], item["sheet"])
            x = item["x"]
            if "series" in item:
                series = [
                    {"name": s["name"], "data": [_sum_range(worksheet, r) for r in s["ranges"]]}
                    for s in item["series"]
                ]
                return {
                    "key": key,
                    "type": item["type"],
                    "title": item["title"],
                    "x": x,
                    "series": series,
                    "error": None,
                }
            y = [_sum_range(worksheet, r) for r in item["ranges"]]
            return {
                "key": key,
                "type": item["type"],
                "title": item["title"],
                "x": x,
                "y": y,
                "error": None,
            }
        elif mode == "header_match":
            df = _read_dataframe(uploads_dir, item["source_file"], item["sheet"])
            x_header = item["x_header"]
            y_header = item["y_header"]
            missing = [h for h in (x_header, y_header) if h not in df.columns]
            if missing:
                raise ValueError(f"找不到表头: '{missing[0]}'")
            x = [_clean_value(v) for v in df[x_header].tolist()]
            y = [_clean_value(v) for v in df[y_header].tolist()]
            return {
                "key": key,
                "type": item["type"],
                "title": item["title"],
                "x": x,
                "y": y,
                "error": None,
            }
        else:
            raise ValueError(f"未知的取数模式: {mode}")
    except Exception as exc:
        return {
            "key": key,
            "type": item.get("type"),
            "title": item.get("title"),
            "x": [],
            "y": [],
            "error": str(exc),
        }


def extract_table(item, uploads_dir):
    key = item["key"]
    try:
        mode = item["mode"]
        if mode == "fixed_range":
            worksheet = _load_worksheet(uploads_dir, item["source_file"], item["sheet"])
            columns = [c["label"] for c in item["columns"]]
            column_values = [_read_fixed_range_column(worksheet, c["range"]) for c in item["columns"]]
            rows = [list(row) for row in zip(*column_values)]
            result = {
                "key": key,
                "title": item["title"],
                "columns": columns,
                "rows": rows,
                "error": None,
                "view_group": item.get("view_group"),
                "view_label": item.get("view_label"),
            }
            if "detail_range" in item:
                result["row_extra"] = _extract_detail_group(worksheet, item["detail_range"])
            return result
        elif mode == "sheet_table":
            columns, rows = _read_sheet_rows(
                uploads_dir, item["source_file"], item["sheet"], item["header_row"], item["id_column"]
            )
            totals = None
            if item.get("show_totals"):
                totals = _build_totals_row(columns, rows, item["id_column"])
            return {
                "key": key,
                "title": item["title"],
                "columns": columns,
                "rows": rows,
                "error": None,
                "view_group": item.get("view_group"),
                "view_label": item.get("view_label"),
                "status_column": item.get("status_column"),
                "totals": totals,
            }
        df = _read_dataframe(uploads_dir, item["source_file"], item["sheet"])
        if mode == "header_match":
            columns = item["columns"]
            missing = [c for c in columns if c not in df.columns]
            if missing:
                raise ValueError(f"找不到表头: '{missing[0]}'")
            rows = [[_clean_value(v) for v in row] for row in df[columns].values.tolist()]
        elif mode == "group_by_sum":
            group_by_header = item["group_by_header"]
            sum_header = item["sum_header"]
            missing = [h for h in (group_by_header, sum_header) if h not in df.columns]
            if missing:
                raise ValueError(f"找不到表头: '{missing[0]}'")
            grouped = df.groupby(group_by_header)[sum_header].sum().sort_values(ascending=False)
            columns = [group_by_header, sum_header]
            rows = [[_clean_value(k), _clean_value(v)] for k, v in grouped.items()]
        else:
            raise ValueError(f"未知的取数模式: {mode}")
        return {
            "key": key,
            "title": item["title"],
            "columns": columns,
            "rows": rows,
            "error": None,
            "view_group": item.get("view_group"),
            "view_label": item.get("view_label"),
        }
    except Exception as exc:
        raw_columns = item.get("columns", [])
        columns = [c["label"] if isinstance(c, dict) else c for c in raw_columns]
        return {
            "key": key,
            "title": item.get("title"),
            "columns": columns,
            "rows": [],
            "error": str(exc),
            "view_group": item.get("view_group"),
            "view_label": item.get("view_label"),
        }


def _resolve_computed_kpi(item, resolved_values):
    key = item["key"]
    label = item["label"]
    operation = item.get("operation")
    if operation != "subtract":
        return {"key": key, "label": label, "value": None, "error": f"未知的运算类型: {operation}"}
    from_key = item.get("from")
    minus_key = item.get("minus")
    from_value = resolved_values.get(from_key)
    minus_value = resolved_values.get(minus_key)
    if from_value is None or minus_value is None:
        missing = from_key if from_value is None else minus_key
        return {"key": key, "label": label, "value": None, "error": f"引用的指标不存在或取值失败: {missing}"}
    return {"key": key, "label": label, "value": from_value - minus_value, "error": None}


def _all_zero_or_none(values):
    return all(v is None or v == 0 for v in values)


def _chart_is_empty(chart):
    if chart.get("error"):
        return False
    if "series" in chart:
        return all(_all_zero_or_none(s["data"]) for s in chart["series"])
    if "y" in chart:
        return _all_zero_or_none(chart["y"])
    return False


def _apply_chart_hint(chart):
    if _chart_is_empty(chart):
        return {
            "key": chart["key"],
            "title": chart.get("title"),
            "hint": "该图表依赖尚未回填的业务字段，暂无数据",
        }
    return chart


def build_dashboard_data(config, uploads_dir):
    resolved_values = {}
    raw_results = {}
    for item in config["kpis"]:
        if item.get("mode") != "computed":
            result = extract_kpi(item, uploads_dir)
            raw_results[item["key"]] = result
            resolved_values[item["key"]] = result["value"]

    kpi_results = []
    for item in config["kpis"]:
        if item.get("mode") == "computed":
            result = _resolve_computed_kpi(item, resolved_values)
        else:
            result = raw_results[item["key"]]
        kpi_results.append(result)

    special_modes = ("shipper_summary", "shipper_summary_monthly")
    batch_source_file = next(
        (t.get("source_file") for t in config["tables"] if t.get("mode") in special_modes),
        None,
    )

    category_sources = {}
    shipper_summary = None
    if batch_source_file:
        try:
            records = _scan_batch_records(uploads_dir, batch_source_file)
            category_sources = build_category_sources(records)
            category_inventory = _read_category_inventory(uploads_dir, batch_source_file)
            shipper_summary = build_shipper_summary(records, category_inventory)
        except Exception:
            shipper_summary = None

    tables = []
    for item in config["tables"]:
        mode = item.get("mode")
        if mode in special_modes:
            key = "by_shipper" if mode == "shipper_summary" else "by_shipper_monthly"
            if shipper_summary is None:
                tables.append({
                    "key": item["key"],
                    "title": item.get("title"),
                    "columns": [],
                    "rows": [],
                    "error": "无法汇总发货单位数据",
                    "view_group": item.get("view_group"),
                    "view_label": item.get("view_label"),
                })
            else:
                data = shipper_summary[key]
                tables.append({
                    "key": item["key"],
                    "title": item["title"],
                    "columns": data["columns"],
                    "rows": data["rows"],
                    "error": None,
                    "view_group": item.get("view_group"),
                    "view_label": item.get("view_label"),
                })
        else:
            tables.append(extract_table(item, uploads_dir))

    return {
        "kpis": kpi_results,
        "charts": [_apply_chart_hint(extract_chart(item, uploads_dir)) for item in config["charts"]],
        "tables": tables,
        "category_sources": category_sources,
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def save_data_json(data, data_path):
    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_data_json(data_path):
    with open(data_path, "r", encoding="utf-8") as f:
        return json.load(f)
